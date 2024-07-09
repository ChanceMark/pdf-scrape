from openpyxl import load_workbook
from pathlib import Path

import pdfplumber
import pandas as pd
import requests

# Load the existing Excel file
existing_excel_path = 'Data.xlsx'
sheet_name = 'Combined x'

try:
    book = load_workbook(existing_excel_path)
except FileNotFoundError:
    print(f"File '{existing_excel_path}' not found.")
    exit(1)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

# Get the sheet by name
sheet = book[sheet_name]

search_header1 = 'Datasheet Link'   # Specify the header value to search for
search_header2 = 'Type'

# Find the column index where the header matches 'search_header'
desc_index = None
type_index = None
row_data = []
colnum = 0
for col in sheet.iter_cols(min_row=2, max_row=2, values_only=True):
    row_data.extend(col)
    for idx, header in enumerate(col, start=1):
        colnum += 1
        
        if header == search_header1:
            desc_index = colnum
        if header == search_header2:
            type_index = colnum

if desc_index is None:
    print(f"Header '{search_header1}' not found in sheet '{sheet_name}'.")
    book.close()
    exit(1)
if type_index is None:
    print(f"Header '{search_header2}' not found in sheet '{sheet_name}'.")
    book.close()
    exit(1)

# Retrieve data from the entire column
column_data = []
for cell in sheet.iter_rows(min_row=3, min_col=desc_index, max_col=desc_index, values_only=True):
    column_data.append(cell[0])

typename = []
for cell in sheet.iter_rows(min_row=3, min_col=type_index, max_col=type_index, values_only=True):
    typename.append(cell[0])


for rownum in range(len(typename)):
# for rownum in range(1):
    
    pdf_document_url = column_data[rownum]
    pdf_document_path = Path(pdf_document_url).name
    print(f"file name: {pdf_document_path}")

    try:
        pdfplumber.open(pdf_document_path)
    except Exception as e:
        response = requests.get(pdf_document_url)
        with open(pdf_document_path, 'wb') as f:
            f.write(response.content)

    dest_row = rownum+3
    dest_col = 0
    flag = []
    desttype = typename[rownum]

    print(f"Type: {desttype}")
    with pdfplumber.open(pdf_document_path) as pdf:
        for i, page in enumerate(pdf.pages):
            pdftables = page.extract_tables()
            if i >1 and i <4:
                for pdftable in pdftables:
                    for pdfdata in pdftable:
                        dest_col = 0
                        for hd in row_data:
                            dest_col += 1
                            if ((hd == pdfdata[0]) or ( (hd == "Rise time 10% - 90% CC" ) and (pdfdata[0] == "Rise time 10 - 90% CC")) or ((hd == "AC input to DC output" ) and (pdfdata[0] == "AC-Input to DC-Output")) 
                                or ((hd == "AC input to case (PE)" ) and (pdfdata[0] == "AC-Input to case (PE)")) or ((hd == "DC output to case (PE)" ) and (pdfdata[0] == "DC-Output to case (PE)"))
                                or ((hd == "DC output to interfaces" ) and (pdfdata[0] == "DC-Output to Interfaces")) or ((hd == "Dimensions (W x H x D)" ) and (pdfdata[0] == "Dimensions (B x H x T)"))) and not(dest_col in flag):
                                flag.append(dest_col)
                                sheet.cell(dest_row, dest_col, pdfdata[1])
                                break
            
            if i == 4 or i==5:            
                for table in pdftables:
                    colnum = 0
                    hcol = 0
                    for hdname in table[0]:
                        colnum +=1
                        if (hdname in desttype) and len(hdname) >2:
                            hcol = colnum-1
                            break
                    
                    if hcol >0 :
                        for pdfdata in table:
                            dest_col = 0
                            for hd in row_data:
                                dest_col += 1
                                if ((hd in pdfdata[0]) or (hd == "Ripple in CV (rms)" and pdfdata[0] == "Ripple rms CV") or (hd == "Ripple in CV (pp)" and pdfdata[0] == "Ripple and noise p-p CV")
                                    or (hd == "Output capacitance" and pdfdata[0] == "Output capacity") or (hd == "Efficiency (up to)" and pdfdata[0] == "Efficiency up to")
                                    or (hd == "Negative DC pole <-> PE" and pdfdata[0] == "Negative DC-Pol <-> PE") or (hd == "Positive DC pole <-> PE" and pdfdata[0] == "Positive DC-Pol <-> PE")
                                    or (hd == "Standard" and pdfdata[0] == "Article number")) and not(dest_col in flag):
                                    # print(f"Equal Header: col:{dest_col}, row:{dest_row}, {pdfdata}")
                                    flag.append(dest_col)
                                    try:
                                        sheet.cell(dest_row, dest_col, pdfdata[hcol])
                                    except Exception as e:
                                        print(f"hcol:{hcol}, pdfdata:{pdfdata}")
                                    break

book.save(existing_excel_path)
book.close()
